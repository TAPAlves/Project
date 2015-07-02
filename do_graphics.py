# -*- coding: utf-8 -*-
"""
Created on Sat Jun 13 20:13:32 2015

@author: Tiago
"""

        
from Bio import Entrez
from do_seqs import do_seqs#classe anterior
Entrez.email = "tiago_alves26@hotmail.com"
import openpyxl
import numpy as np
import matplotlib.pyplot as plot

class do_graphics:

    def __init__(self):
        self.media={}
        self.sd={}
        self.results={}
        self.proj=do_seqs()
        self.mutationsPos=[]
        
        
        
    def create_mutations(self,pos,am_init,am_fin):
        '''
        Funcao que dadas listas de posicoes, aminoacido inicial e final, 
        as insere na classe para posterior adicao a sequencia normal na classe do_seqs
        '''        
        dic={}
        for p in range(len(pos)):
            dic[p[pos]]=(am_init[pos],am_fin[pos])#obriga a que as listas sejam dadas no mesmo tamanho
        for key in dic.keys():
            self.proj.mutations[key]=dic[key]
            
    
    def import_results(self,fnormal,fmutated):
        '''
        Funcao que importa os ficheiros com os resultados das análises do NETMHC(I/II)PAN
        '''
        mhc=[]
        mhc_m=[]
        wb_norm=openpyxl.load_workbook(str(fnormal))#abre o ficheiro com os resultados da sequencia normal
        sheets=wb_norm.get_sheet_names()#obtem-se o nome das folhas
        sheet_norm=wb_norm.get_sheet_by_name(sheets[0])#primeira folha- e unica
        #print sheet_norm.calculate_dimension()#calcular dimensoes A1:AF1624
        #print len(sheet_norm.rows[1])#tamanho da linha
        mhc_mol_line=sheet_norm.rows[0]#tuplo com celulas desta linha
        for mhc_mol in mhc_mol_line:#cada celula
            if mhc_mol.value is not None:#celulas com valor
                mhc.append(mhc_mol.value)
        pos=self.proj.mutations.keys()#vai buscar a classe anterior as posicoes das mutacoes
        #########pos=[132]
        for p in pos:
            search=sheet_norm.columns[0]
            for cell in range(len(search)):
                if search[cell].value==int(p)-1:
                    tam=len(sheet_norm.cell(None,cell,1).value)#para retirar o tamanho da proteina respetiva
                    if tam not in self.results:
                        self.results[tam]={}
                        self.results[tam]['Normal']={}
                    else:
                        if 'Normal' not in self.results[tam]:
                            self.results[tam]['Normal']={}
                    for hla in mhc:
                        if hla not in self.results[tam]['Normal']:
                            self.results[tam]['Normal'][hla]=[]
                        mhc_mol_line=sheet_norm.rows[0]#tuplo com celulas desta linha
                        for mhc_mol in range(len(mhc_mol_line)):#cada celula. ou seja, cada coluna
                            if mhc_mol_line[mhc_mol].value==hla:
                                if int(p)-tam>=0:
                                    for val in range(int(p)-tam,int(p)):
                                        valores=str(sheet_norm.cell(None,val+2,mhc_mol).value)#valores finais pretendidos
                                        valores_finais=valores.replace('.', '')
                                        self.results[tam]['Normal'][hla].append(int(valores_finais))       
                                else:
                                    for val in range(0,int(p)):
                                        valores=str(sheet_norm.cell(None,val+2,mhc_mol).value)#valores finais pretendidos
                                        valores_finais=valores.replace('.', '')
                                        self.results[tam]['Normal'][hla].append(int(valores_finais))  
          
        wb_mut=openpyxl.load_workbook(str(fmutated))
        sheets_m=wb_mut.get_sheet_names()
        sheet_mut=wb_mut.get_sheet_by_name(sheets_m[0])#primeira folha- e unica
        mhc_mol_line_m=sheet_mut.rows[0]#tuplo com celulas desta linha
        for mhc_mol in mhc_mol_line_m:#cada celula
            if mhc_mol.value is not None:#celulas com valor
                mhc_m.append(mhc_mol.value)

        for p in pos:
            search=sheet_mut.columns[0]
            for cell in range(len(search)):
                if search[cell].value==int(p)-1:
                    tam=len(sheet_mut.cell(None,cell,1).value)#para retirar o tamanho da proteina respetiva
                    if tam not in self.results:
                        self.results[tam]={}
                        self.results[tam]['Mutated']={}
                    else:
                        if 'Mutated' not in self.results[tam]:
                            self.results[tam]['Mutated']={}

                    for hla in mhc_m:
                        if hla not in self.results[tam]['Mutated']:
                            self.results[tam]['Mutated'][hla]=[]
                        mhc_mol_line=sheet_mut.rows[0]#tuplo com celulas desta linha
                        for mhc_mol in range(len(mhc_mol_line)):#cada celula. ou seja, cada coluna
                            if mhc_mol_line[mhc_mol].value==hla:
                                if int(p)-tam>=0:
                                    for val in range(int(p)-tam,int(p)):
                                        valores=str(sheet_mut.cell(None,val+2,mhc_mol).value)#valores finais pretendidos
                                        valores_finais=valores.replace('.', '')
                                        self.results[tam]['Mutated'][hla].append(int(valores_finais))       
                                else:
                                    for val in range(0,int(p)):
                                        valores=str(sheet_mut.cell(None,val+2,mhc_mol).value)#valores finais pretendidos
                                        valores_finais=valores.replace('.', '')
                                        self.results[tam]['Mutated'][hla].append(int(valores_finais))  
        
               
               
    def do_mean(self):#tem de se dar o inteiro
        '''
        Funcao que dados os resultados importados, calcula a sua media
        e retorna os dicionarios já existentes mas apenas com o valor da media dos valores de entrada
        '''
        memory=[]
        self.media=self.results.copy()#criar copia de dicionario
        for pos in self.media.keys():#posicoes
            for state in self.media[pos]:#normal e mutado
                for hla in self.media[pos][state]:#hla
                    soma=0
                    num=0                    
                    for i in range(len(self.media[pos][state][hla])):
                        soma+=self.media[pos][state][hla][i]#soma dos elementos da lista de valores                        
                        num+=1                        
                    
                    memory.append(float(soma/num))
                    self.media[pos][state][hla]=[]#limpar a lista
                    self.media[pos][state][hla].append(memory[0])
                    memory=[]#limpar a memoria
                    
    
    def do_sd(self):#tem de se dar o inteiro
        '''
        Funcao que dados os resultados importados, calcula o seu desvio padrão
        e retorna os dicionarios já existentes mas apenas com o valor do desvio dos valores de entrada
        '''
        memory=[]
        hlavalue=[]
        self.sd=self.results.copy()#criar copia de dicionario
        for pos in self.sd.keys():#posicoes
            for state in self.sd[pos]:#normal e mutado
                for hla in self.sd[pos][state]:#hla
                    for i in range(len(self.sd[pos][state][hla])):
                        hlavalue.append(self.sd[pos][state][hla][i])  
                    sd=np.std(hlavalue)
                    memory.append(int(sd))
                    self.sd[pos][state][hla]=[]#limpar a lista
                    self.sd[pos][state][hla].append(memory[0])
                    memory=[]#limpar a memoria
                    hlavalue=[]
                
    
    def variables_for_graph(self,pos):
        '''
        Funcao que retorna as variaveis a ser usadas pelo construtor do grafico de barras
        '''
        N = 0 #depende dos hlas que existirem em análise
        maximo=0#eixo y
        hlaList=[]#hla para legenda
        normal=[]#valores normais
        mutated=[]#valores mutados       
        for state in self.media[pos].keys():#normal e mutado
            if len(hlaList)==0:
                for hla in self.media[pos][state].keys():#
                    N+=1
                    hlaList.append(hla)#lista com hlas
            else:
                pass#só interessa uma corrida (um deles mutado ou nao preenche os requisitos hla e num)
        for hla in hlaList:#para garantir que os valores sao correspondentes nos dois caso porque a ordem podia ser diferente
            valorn=int(self.media[pos]['Normal'][hla][0])
            valorm=int(self.media[pos]['Mutated'][hla][0])
            normal.append(valorn)
            mutated.append(valorm)
                
        #calculo do maximo para atribuir ao eixo y
        if max(normal)>max(mutated):
            maximo=max(normal)
            percent=0.25*maximo
            maximo+=percent
        else:
            maximo=max(mutated)
            percent=0.25*maximo
            maximo+=percent

        return N, maximo,hlaList, normal, mutated#todos os parametros necessários para a construção do grafico   

        
            
    def do_graphics(self):
        '''
        Funcao que faz o plot dos graficos da afinidade media dos peptidos 
        com a mutacao vs diferentes HLA classe I ou II em estudo 
        (numero de graficos depende dos diferentes tamanho das proteinas)
        '''
        for pos in self.media.keys():
            figure = plot.figure()
            subplot = figure.add_subplot(111)
            N,maximo,hlaList,normal,mutated=self.variables_for_graph(pos)
            
            #variaveis necessarias
            ind = np.arange(N)# localizacoes nos grupos
            width=0.35 #largura das barras
            ## the bars
            bar1 = subplot.bar(ind, normal, width,color='lightblue')
                               #yerr=stdesvn
                               #,error_kw=dict(elinewidth=9,ecolor='black'))
            bar2 = subplot.bar(ind+width, mutated, width,color='darkred')
                               #yerr=stdesvm)
                              #error_kw=dict(elinewidth=2,ecolor='black'))
            # axes and labels4
            subplot.set_xlim(-width,len(ind)+width)#tamanho do eixo do x                                
            subplot.set_ylim(0,maximo)#calcular o valor maior para colocar aqui (pela função)
            subplot.set_ylabel('HLA binding affinity (nM)')
            title='%s-mer peptides affinity to HLA'%pos            
            subplot.set_title(title)# colocar aqui o tamanho (isto tem de estar num ciclo)
            xNames = [i for i in hlaList]#Todos os HLAs
            subplot.set_xticks(ind+width)#posição da legenda
            xLegend = subplot.set_xticklabels(xNames)
            plot.setp(xLegend, rotation=30, fontsize=10)
            if xNames[0][0]=='H':
                xl=subplot.set_xlabel('HLAs class I')
            else:
                xl=subplot.set_xlabel('HLA class II')
            plot.setp(xl,rotation=0,fontsize=10)
            ## add a legend
            subplot.legend( (bar1[0], bar2[0]), ('WildType', 'Mutated') )
            plot.show()
        
                    
                    
###############################################################################
################################  M  E  N  U  #################################         
###############################################################################        
       
def main():
    p=do_graphics()   
    while True:
        #print ("\nRESPONDA SEMPRE DENTRO DE ASPAS, A EXCEPÇÂO DA SELECAO DO MENU (INTEIROS)\n\n" )
        print ('\n##### PHASE I - Extraction and creation of sequences (WildType and Mutated) ######\n')        
        print("1 - Get a 'fasta' file with gene sequence of NCBI database")
        print("2 - Translate DNA sequence to a protein for application on NetMHCpan or NetMHCIIpan servers")
        print("3 - Search for protein id's on NCBI")
        print("4 - Get a 'fasta' file with protein sequence of NCBI database")
        print("5 - Insert mutations on a protein sequence, getting the 'fasta' file") 
        print("6 - Exit")        
        
        print ("\n##### PHASE II - NetMHCpan or NetMHCIIpan results analysis ######\n")
        print("7 - Print a barplot with obtained results of NetMHCpan or NetMHCIIpan for Wildtype and Normal sequences")
        print("8 - Exit")        
        op=raw_input("What do you want to do? ")
        if op=='1':
            gi=raw_input('\nDo you have GI (Gene database)? (Y or N)\n').upper()
            if gi=='Y':
                num=raw_input('\nInsert GI number:\n')
                a=p.proj.get_seq_from_gene(0,num)
            elif gi=='N':
                term=raw_input('\nInsert a gene name:\n')
                a=p.proj.get_seq_from_gene(term)
            else:
                print ('Invalid input! Try again.')
        elif op=='2':
            ask=raw_input("\nDo you already run option 1 (Y or N)").upper()
            if ask=='Y':
                try:
                    p.proj.convert_into_protein(a)
                except:
                    print ('Option 1 not runned! Please repeat option 1!')
            elif ask=='N':
                gi=raw_input('\nDo you have GI (Gene database)? (Y or N)\n').upper()
                if gi=='Y':
                    num=raw_input('\nInsert GI number:\n')
                    a=p.proj.get_seq_from_gene(0,num)
                    p.proj.convert_into_protein(a)
                elif gi=='N':
                    term=raw_input('\nInsert a gene name:\n')
                    a=p.proj.get_seq_from_gene(term)
                    p.proj.convert_into_protein(a)
                else:
                    print ('Invalid input! Try again.') 
                
            else:
                print ('Invalid input! Try again.')            
        
        elif op=='3':
            term=raw_input('\nInsert a protein name:\n')
            print ("\nGI's:\n"+str(p.proj.search_seq_ids(term)))
             
        
        elif op=='4':
            gi=raw_input('\nDo you have GI (Gene database)? (Y or N)\n').upper()
            if gi=='Y':
                num=raw_input('\nInsert GI number:\n')
                p.proj.create_file_with_seq(num)
            elif gi=='N':
                ask=raw_input('\nDo you want to search for protein GIs (ID) or do you accept the first entry of obtained results (FIRST)? (ID or FIRST)\n').upper()
                if ask=='ID':
                    ask2=raw_input('\nInsert a protein name:\n')
                    print ("\nGI's:\n"+str(p.proj.search_seq_ids(ask2)))
                    ask3=raw_input('\nInsert GI number:\n')
                    p.proj.create_file_with_seq(ask3)
                elif ask=='FIRST':
                    ask2=raw_input('\nInsert a protein name:\n')
                    p.proj.create_file_with_seq(0,ask2)
                else:
                    print ('Invalid input! Try again.')
            else:
                print ('Invalid input! Try again.')
        
                    
        elif op=='5':
            ask=raw_input("\nDo you already run option 2 or option 4 (Y or N)\n").upper()
            if ask=='Y':
                try:
                    a=raw_input('\nInsert the mutation position(s):(example: 1,2,3)\n')
                    b=raw_input('\nInsert the initial aminoacid for each position above(with the same order):(example: A,B,C)\n').upper()
                    c=raw_input('\nInsert the final aminoacid for each position above(with the same order):(example: A,B,C)\n').upper()
                    a1=[]
                    b1=[]
                    c1=[]
                    for pos in a.split(','):
                        a1.append(pos)
                    for pos in b.split(','):
                        b1.append(pos)
                    for pos in c.split(','):
                        c1.append(pos)
                    p.proj.create_mutations(a1,b1,c1)
                    p.proj.induce_mutation()
                    p.proj.create_file_with_seq('',0,mutated=True)
                except:
                    print ('Option 2/option 4 not runned! Please repeat one of the options!')
                    
            elif ask=='N':
                print ("\n**Obtention of 'fasta' file with protein sequence**\n")
                gi=raw_input('\nDo you have GI (Gene database)? (Y or N)\n').upper()
                if gi=='Y':
                    num=raw_input('\nInsert GI number:\n')
                    p.proj.create_file_with_seq(num)
                    a=raw_input('\nInsert the mutation position(s):(example: 1,2,3)\n')
                    b=raw_input('\nInsert the initial aminoacid for each position above(with the same order):(example: A,B,C)\n').upper()
                    c=raw_input('\nInsert the final aminoacid for each position above(with the same order):(example: A,B,C)\n').upper()
                    a1=[]
                    b1=[]
                    c1=[]
                    for pos in a.split(','):
                        a1.append(pos)
                    for pos in b.split(','):
                        b1.append(pos)
                    for pos in c.split(','):
                        c1.append(pos)
                    p.proj.create_mutations(a1,b1,c1)
                    p.proj.induce_mutation()
                    p.proj.create_file_with_seq(num,0,mutated=True)
                elif gi=='N':
                    ask=raw_input('\nDo you want to search for protein GIs (ID) or do you accept the first entry of obtained results (FIRST)? (ID or FIRST)\n').upper()
                    if ask=='ID':
                        ask2=raw_input('\nInsert a protein name:\n')
                        print ("\nGI's:\n"+p.proj.search_seq_ids(ask2))
                        ask3=raw_input('\nInsert GI number:\n')
                        p.proj.create_file_with_seq(ask3)
                        a=raw_input('\nInsert the mutation position(s):(example: 1,2,3)\n')
                        b=raw_input('\nInsert the initial aminoacid for each position above(with the same order):(example: A,B,C)\n').upper()
                        c=raw_input('\nInsert the final aminoacid for each position above(with the same order):(example: A,B,C)\n').upper()
                        a1=[]
                        b1=[]
                        c1=[]
                        for pos in a.split(','):
                            a1.append(pos)
                        for pos in b.split(','):
                            b1.append(pos)
                        for pos in c.split(','):
                            c1.append(pos)
                        p.proj.create_mutations(a1,b1,c1)
                        p.proj.induce_mutation()
                        p.proj.create_file_with_seq(ask3,0,mutated=True)#proteina criada acima é a normal e esta é a mutada
                        
                    elif ask=='FIRST':
                        ask2=raw_input('\nInsert a protein name:\n')
                        p.proj.create_file_with_seq(0,ask2)
                        a=raw_input('\nInsert the mutation position(s):(example: 1,2,3)\n')
                        b=raw_input('\nInsert the initial aminoacid for each position above(with the same order):(example: A,B,C)\n').upper()
                        c=raw_input('\nInsert the final aminoacid for each position above(with the same order):(example: A,B,C)\n').upper()
                        a1=[]
                        b1=[]
                        c1=[]
                        for pos in a.split(','):
                            a1.append(pos)
                        for pos in b.split(','):
                            b1.append(pos)
                        for pos in c.split(','):
                            c1.append(pos)
                        p.proj.create_mutations(a1,b1,c1)
                        p.proj.induce_mutation()
                        p.proj.create_file_with_seq(0,ask2,mutated=True)
                    else:
                        print ('Invalid input! Try again.')
                else:
                    print ('Invalid input! Try again.')
            else:
                print ('Invalid input! Try again.')
                
        elif op=='7':
            if len(p.proj.mutations.keys())==0:#Nao se efetuou a primeira fase de elaboracao das sequencias
                print ('\nPhase I not runned!!\n')
                a=raw_input('\nInsert the mutation position(s)(example: 1,2,3):\n')
                a1=[]
                b1=[]
                c1=[]
                for pos in a.split(','):
                    a1.append(pos)
                    b1.append('A')#so interessa saber as posicoes nesta fase
                    c1.append('A')
                p.proj.create_mutations(a1,b1,c1)
                fn=raw_input('\nInsert the name of the WildType file (example : abcd.xlsx):\n')
                fm=raw_input('\nInsert the name of the Mutated file (example : abcd.xlsx):\n')
                p.import_results(fn,fm)#importar os resultados
            else:
                fn=raw_input('\nInsert the name of the WildType file (example : abcd.xlsx):\n')
                fm=raw_input('\nInsert the name of the Mutated file (example : abcd.xlsx):\n')
                p.import_results(fn,fm)#importar os resultados
            p.do_mean()#faz a media
            p.do_graphics()#faz os graficos
            break
            
            
        elif op=='6' or op=='8':
            print("\n Bye!")
            break
        
        else:
            print("\n This option in not valid! Try again.")
        
       
         
       
if __name__=='__main__':
    
    def test1():
        p=do_graphics()
        #p.variables_for_graph(8)
        #print (p.import_results('23780_NetMHCpan_NORMAL_STATE.xls','0'))
        print p.import_results('132normalI.xlsx', '132mutadoI.xlsx')
        p.do_mean()
        #print p.media
        #print p.sd        
        #p.do_graphics()
        #p.do_sd()
        #p.do_mean()
        #print p.variables_for_graph(8)
        p.do_graphics()
    #test1()
    
    main()
        
            